# UI
# Author :      Nathan Krueger
# Created       5:00 PM 7/16/15
# Last Updated  4:00 PM 12/2/16
# Version       2.75

import excel
from openpyxl import *
from os.path import exists

corpora = []
max_index = 0
output = 0
sort_type = 0
dual = False

menu = """
Enter a command:
output   change type of output between shell text and external documentation
sort     sort the output based on a particular value
newc     add a new corpus
listc    lists all available corpora by registered name
swa      single word analysis
swac     single word analysis with a corpus
mwaw     multi-word analysis using word definitions and POS
mwax     multi-word analysis using excel and wordnet data
mwac     multi-word analysis with a corpus
dwsa     dual-word similarity analysis
polys    polysemy rating for a selection of words
mindep   mindepth of word/s (first result only, see dtree for alternate synsets)
pol_min  runs both polysemy and mindepth analysis of a selection of words
dtree    depth tree of a given word (working on making it neater)
xhyper   returns the highest order x (given number) hypernyms of a word or words

q        quit

command: """

#experimental shortcut:
"""def pos(data)->[str]:
    '''returns a list of strings corresponding to the appropriate part of speech'''
    result = []
    for def_index in range(len(data[1])):
        if data[2][def_index] == 'n':
            result.append("noun")
        if data[2][def_index] == 'a':
            result.append("adjective")
        if data[2][def_index] == 's':
            result.append("satellite adjective")
        if data[2][def_index] == 'r':
            result.append("adverb")
        if data[2][def_index] == 'v':
            result.append("verb")
    return result"""

def setup()->None:
    '''sets up the user interface'''
    global corpora, max_index
    corpora = open('corpora.txt').read().splitlines()
    max_index = len(corpora)
    return

def interface()->(str, int):
    """asks the user what to do and then asks the controller to do it"""
    global output, sort_type, dual
    dual = False
    while True:
        cmd = input(menu).strip().lower()
        if cmd == 'output':
            output = change_output()
            return (None,None)
        elif cmd == 'sort':
            sort_type = sort_change()
            return (None,None)
        elif cmd == 'newc':
            return ('newc', newc())
        elif cmd == 'listc':
            print('Available corpora:')
            for c in corpora:
                print(c)
            return (None,None)
        elif cmd == 'swa':
            return ('swa', swa())
        elif cmd == 'swac':
            return ('swac', swac())
        elif cmd == 'mwaw':
            return ('mwaw', mwa())
        elif cmd == 'mwax':
            return ('mwax', mwa())
        elif cmd == 'mwac':
            return ('mwac', mwac())
        elif cmd == 'dwsa':
            dual = True
            return ('dwsa', mwa())
        elif cmd == 'polys':
            return ('polys', mwa())
        elif cmd == 'mindep':
            return ('mindep', mwa())
        elif cmd == 'pol_min':
            return ('pol_min', mwa())
        elif cmd == 'dtree':
            return ('dtree', swa())
        elif cmd == 'xhyper':
            return ('xhyper', mwa())
        elif cmd == 'q':
            print("Goodbye")
            return ('quit', None)
        else:
            print("Invalid command, please try again")
    return

def change_output()->int:
    '''allows user to change the programs output to text and/or excel file'''
    while True:
        output_type = input("Change output to (excel, default/text/shell text, both): ").strip().lower()
        if output_type in ['text', 'shell text', 'default']:
            return 0
        elif output_type == 'excel':
            return 1
        elif output_type == 'both':
            return 2
        else:
            print("Invalid output type, please select from: excel, text, or both")

def sort_change()->str:
    '''changes how the outputs are ordered'''
    print("The words will be printed in order followed by their analysis")
    possibles = ['default','def_count','alphabetical','length']#'TASA','AWL','sfi (from Zeno)','d (from Zeno)']
    print("Possible sort methods: {}".format(possibles))
    while True:
        result = input("Please enter the basis for output sorting: ").strip().lower()
        if result not in possibles:
            print("This sort method is not valid")
        break
    if result == 'default':
        return 0
    elif result == 'def_count':
        return 1
    elif result == 'alphabetical':
        return 2
    elif result == 'length':
        return 3

def newc()->('file', str):
    """tell controller to add a new corpus given its name"""
    print("""Please enter the fileid of the new corpus exactly as it is
(including the extension; only .txt is currently supported) and that the
file is present in the current directory:
""")
    while True:
        file_name = input("fileid of new corpus: ")
        try:
            if file_name.split('.')[1] != 'txt':
                print("This file format is not currently supported")
                continue
            assert exists(file_name)
        except:
            print("This corpus is not available, please make sure that you typed it correctly")
            return newc()
        name = input("Please enter the name of the corpus: ")
        break
    return (file_name, name)

def swa()->str:
    """run an analysis on a signle word"""
    return input("Please enter a word to analyze: ").strip().lower()

def swac()->(int,str):
    """run an analysis on a signle word based on a corpus"""
    print('Available corpora:')
    for c in corpora:
        print(c)
    while True:
        corpus = input("Please enter an available corpus to refrence by index number: ")
        try:
            corpus = int(corpus)
            if corpus > max_index or corpus < 1:
                print("This corpus does not exist or is not available")
            else:
                break
        except:
            print("This is not a valid corpus index (valid indicies are 1-{})".format(max_index))
    word = input("Please enter a word to analyze: ").strip().lower()
    return (corpus, word)

def request_x()->int:
    """asks user to provide an integer representing the number of hypernyms to return from controller"""
    while True:
        try:
            x = int(input("Please enter an integer, x, representing how many high order hypernyms\nto return per word: ").strip().lower())
            break
        except:
            print("X must be an integer")
    return x

def mwa()->[str]:
    """run an analysis on several words"""
    while True:
        print("Word sources: text, excel, manual (.txt, .xls/.xlsx, type into console)")
        word_source = input("Please enter a source for the words to analyze: ").strip().lower()
        if word_source not in ['text', 'excel', 'manual']:
            print("This is not a valid word source")
            continue
        break
    if word_source == 'manual':
        if not dual:
            words = input("Please enter a series of words to analyze seperated only by spaces: ").strip().lower().split()
        else:
            while(True):
                words = input("Please enter pairs of words to analyze seperated only by spaces: ").strip().lower().split()
                if len(words)%2 != 0:
                    print("There an even number of words for this function")
                else:
                    break
        return words
    if word_source == 'excel':
        return excel_words()
    if word_source == 'text':
        return text_words()

def text_words()->[str]:
    '''takes a text file and returns a list of words'''
    print("""Please enter the fileid of the text file exactly as it is
(including the extension), please ensure that the file is present in
the current folder and that all words seperated by only whitespace
""")
    while True:
        file_name = input("fileid of text file: ").strip()
        try:
            file = open(file_name).read().splitlines()
            break
        except:
            print("This text file is not available, please make sure that you typed it correctly")
            continue
    result = []
    for line in file:
        line = line.strip().lower()
    if len(file[0]) == 1:
        result = file
        return result
    for line in file:
        for word in line.split():
            result.append(word)
    return result

def excel_words()->[str]:
    """run an analysis on many words from an excel file"""
    print("""Please enter the fileid of the excel file exactly as it is
(including the extension), please ensure that the file is present in
the current folder and that all words are listed in the first column
""")
    while True:
        file = input("fileid of excel document: ").strip()
        sheet = input("Please enter the sheet name to check (by default, Excel uses Sheet1): ").strip()
        try:
            file = load_workbook(file)
            sheet = file.get_sheet_by_name(sheet)
            break
        except:
            print("This excel document or sheet is not available, please make sure that you \ntyped it correctly")
            continue
    result = []
    for pos in range(len(tuple(sheet.rows))):
        try:
            result.append(str(sheet.cell(row = pos + 1, column = 1).value))
            if dual:
                result.append(str(sheet.cell(row = pos + 1, column = 2).value))
        except:
            pass
    return result

def mwac()->(int,[str]):
    """run an analysis on several words based on a corpus"""
    print('Available corpora:')
    for c in corpora:
        print(c)
    while True:
        corpus = input("Please enter an available corpus to refrence by index number: ")
        try:
            corpus = int(corpus)
            if corpus > max_index or corpus < 1:
                print("This corpus does not exist or is not available")
            else:
                break
        except:
            print("This is not a valid corpus index (valid indicies are 1-{})".format(max_index))
    return (corpus, mwa())

def output_data(value)->None:
    '''selects between output styles and call the correct function/s'''
    value = (sort(value[0]), value[1])
    if output != 1:
        print_data(value)
    if output != 0:
        data_to_file(value)

def data_to_file(value)->None:
    '''outputs the data as a file of specified name and type'''
    if value[1] in ['newc','dtree']:
        print("Error: newc and dtree do not support Excel output as their outputs\nare variably sized or empty")
        return
    print("""
Warning: for the moment this program cannot append to files,
only create and overwrite them, please be careful about using existing files
You can specify a different sheet name to use if you wish to add to a file

Also, please choose a file type of .xlsx (.xls is not currently supported)
and be sure that the file is not open in another window.
          """)
    while True:
        file_name = input("Please enter the name of the file you would to create, including the extension: ")
        print("\n{}\n".format(file_name))
        sure = input("Are you sure this is the file name you wish to use (y/n)").strip().lower()
        if sure not in ['y', 'yes']:
            continue
        if file_name.split('.')[1] != 'xlsx':
            print("This file type is not currently supported")
            continue
        sheet_name = input("Please enter a sheet name: ").strip()
        break
    file_from_data(value, file_name, sheet_name)
    return

def file_from_data(value, file_name: str, sheet_index: int)->None:
    '''creates an output file using the given name'''
    excel.file_setup(value, file_name, sheet_index)
    return

def print_data(value)->None:
    """prints the data"""
    #can probably shorten this if I try
    data, function = value
    if function == 'newc':
        if data:
            print("New corpus has been sucesfully installed")
        else:
            print("New corpus could not be installed...")
    if function in ['swa', 'swac']:
        print("\nWord: {}".format(data[0][0]))
        if function == 'swac':
            print_nltk(data[0][1])
        print_wordnet(data[0][2])
        print_excel(data[0][3])
    if function in ['mwax', 'mwac']:
        for word in data:
            print("\n{}\nWord: {}".format('*'*80,word[0]))
            if function == 'mwac':
                print_nltk(word[1])
            print_wordnet(word[2])
            print_excel(word[3])
    if function in ['mwaw']:
        for word in data:
            print("\n{}\nWord: {}".format('*'*80,word[0]))
            print_word_def(word[1:])
    if function == 'dwsa':
        #weird things here, check what happens when result is none
        print("\nWord 1                Word 2\nSimilarity (LCH, WUP, Path)")
        for word in data:
           print("\n{:18}{:18}\n{}   {}   {}".format(word[0],word[1],word[2],word[3],word[4])) 
    if function == 'polys' or function == 'pol_min':
        print("\nNumber of defintions per part of speech for each word:")
        print("\nWord                Noun  Adj  SatAdj  Adv  Verb")
        for word in data:
            print("{:18}{:6}{:5}{:8}{:5}{:6}".format(word[0],word[1],word[2],word[3],word[4],word[5]))
    if function == 'pol_min':
        print("""\nMin depth of the first, or most common definition for each part of
speech of each word; -1 signifies that no defintion of that type was found:""")
        print("\nWord  Min depth as: Noun  Adj  SatAdj  Adv  Verb")
        for word in data:
            print("{:18}{:6}{:5}{:8}{:5}{:6}".format(word[0],word[6],word[7],word[8],word[9],word[10]))
    if function == 'mindep':
        print("""\nMin depth of the first, or most common definition for each part of
speech of each word; -1 signifies that no defintion of that type was found:""")
        print("\nWord  Min depth as: Noun  Adj  SatAdj  Adv  Verb")
        for word in data:
            print("{:18}{:6}{:5}{:8}{:5}{:6}".format(word[0],word[6],word[7],word[8],word[9],word[10]))

    if function == 'dtree':
        print_dtree(data)
    if function == 'xhyper':
        print_xhyper(data)
    return

def print_nltk(data)->None:
    '''prints data from corpus'''
    print("""
Total number of tokens:             {}
Number of unique tokens:            {}
Richness of the text:               {}
Count of word's occurences:         {}
Rate of word's occurence per token: {}
""".format(data[0],data[1],data[2],data[3], data[4]))
    return

def print_wordnet(data)->None:
    '''prints data from wordnet'''
    #parts of speech and defintions
    print("Defintions:")
    for def_index in range(len(data[0])):
          if data[1][def_index] == 'n':
              print("noun: {}".format(data[0][def_index]))
          if data[1][def_index] == 'a':
              print("adjective: {}".format(data[0][def_index]))
          if data[1][def_index] == 's':
              print("satellite adjective: {}".format(data[0][def_index]))
          if data[1][def_index] == 'r':
              print("adverb: {}".format(data[0][def_index]))
          if data[1][def_index] == 'v':
              print("verb: {}".format(data[0][def_index]))
    #related words
    print("\nRelated words:")
    print("Synonyms:")
    for syn in data[2]:
        print(syn)
    print("\nAntonyms:")
    for ant in data[3]:
        print(ant)
    print("\nHypernyms:")
    for hyper in data[4]:
        print(hyper)
    print("\nHyponyms:")
    for hypo in data[5]:
        print(hypo)
    return

def print_word_def(data)->None:
    '''prints a words definitions, POS, etc'''
    print("Definition # Synset                Depth POS    POS definition # Definition")
    for num in range(len(data)):
        print("{:3} {:30} {:5} {:19} {:3} {}".format(data[num][1], str(data[num][2]), data[num][3], data[num][4], data[num][5], data[num][6]))
    return

def print_excel(data)->None:
    '''prints data from excel spreadsheets'''
    print("\nTASA number: {}".format(data[2]))
    
    print("\nAOA data:")
    print("OccurTotal:      {}\nOccurNum:        {}\nFreq_pm:         {}\nRating.Mean:     {}\nRating.SD:       {}\n(unknown value): {}".format(
        data[1][0],data[1][1],data[1][2],data[1][3],data[1][4],data[1][5]))
    
    print("\nAWL value: {}".format(data[0]))
    
    print("\nSUBTLEX data:")
    print("""FREQcount:  {}\nCScount:    {}\nFREQlow:    {}\nCDlow:      {}\nSUBTL_WF:   {}\nLog_10(WF): {}
SUBTL_CD:   {}\nLog_10(CD): {}""".format(data[3][0],data[3][1],data[3][2],data[3][3],
                                       data[3][4],data[3][5],data[3][6],data[3][7]))
    
    print("\nZeno data:")
    print("""sfi:  {}\nd:    {}\nu:    {}\nf:    {}\ngr1:  {}\ngr2:  {}\ngr3:  {}\ngr4:  {}\ngr5:  {}\ngr6:  {}
gr7:  {}\ngr8:  {}\ngr9:  {}\ngr10: {}\ngr11: {}\ngr12: {}\ngr13: {}""".format(
    data[4][0],data[4][1],data[4][2],data[4][3],data[4][4],data[4][5],data[4][6],
    data[4][7],data[4][8],data[4][9],data[4][10],data[4][11],data[4][12],data[4][13],
    data[4][14],data[4][15],data[4][16]))
    return

def print_dtree(data):
    '''prints dtree data'''
    from pprint import pprint
    depth = lambda L: isinstance(L, list) and max(map(depth, L))+1
    d_count = 0
    if (len(data[1]) > 1):
        print("Multiple definitions for the word {} detected: \n".format(data[0]))
        print("Defintions:")
        for def_index in range(len(data[1])):
            if data[2][def_index] == 'n':
                print("{}: noun: {}".format(def_index, data[1][def_index]))
                d_count = d_count + 1
        while True:
            action = input('\nWhich dtree would you like (choose an index or say "all"): ')
            if action == 'all':
                print("\nNote: multiple entires on the same line are equivalent")
                for def_index in range(d_count):
                    print("\nDtree of '{}' with depth {} as defined as: {}".format(data[0],depth(data[3][def_index]) - 1,data[1][def_index]))
                    pprint(data[3][def_index])
                return
            else:
                try:
                    action = int(action)
                    if (action >= d_count):
                        raise fail
                    dtree = data[3][int(action)]
                    #in this order to make the note only appear if it works
                    print("\nNote: multiple entires on the same line are equivalent")
                    print("\nDtree of '{}' with depth {} as defined as: {}".format(data[0],depth(dtree) - 1,data[1][action]))
                    pprint(dtree)
                    return
                except:
                    print('This index is either invalid or not "all"')
    elif (len(data[1]) == 1):
        print('Only 1 defintion of the word "{}" wss found: '.format(data[0]))
        print("\nNote: multiple entires on the same line are equivalent")
        print("\nDtree of '{}' as defined as: {}".format(data[0],data[1][0]))
        pprint(data[3][0])
    else:
        print('No defintions for the word "{}" were found...'.format(data[0]))
    return

def print_xhyper(data):
    """prints the data from xhyper"""
    for word in data[1:]:
        print("Word: {}\nSynset used: {}\nPOS: {}".format(word[0], word[2], word[1]))
        for x in word[3]:
            if x is None:
                break
            print("\t{}".format(x))
    return

def sort(data)->list:
    '''uses the current sort method to sort the data'''
    if sort_type == 0:
        return data
    if sort_type == 1:
        data.sort(key = lambda x: len(x[2][0]), reverse = True)
    if sort_type == 2:
        data.sort(key = lambda x: x[0])
    if sort_type == 3:
        data.sort(key = lambda x: len(x[0]))
    return data

#parts of speech conversion: ADJ, ADJECTIVE_SATELLITE, ADV, NOUN, VERB = 'a', 's', 'r', 'n', 'v'
#for synset in wn.synsets('mint', wn.NOUN):
#     print(synset.name() + ':', synset.definition())


#if __name__ == '__main__':
    #interface()
