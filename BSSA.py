# Binary Synset Similarity Analysis
# Author :      Nathan Krueger
# Created       6:00 PM 2/26/16
# Last Updated  5:00 PM 3/06/16
# Version       1.0

import nltk
import excel
from openpyxl import *
from nltk.corpus import wordnet


def excel_file(file_name):
    '''Asks for a sheet name and reads from the given excel file'''
    while True:
        sheet = input("Please enter the sheet name to check (by default, Excel uses Sheet1): ").strip()
        try:
            file = load_workbook(file_name)
            sheet = file.get_sheet_by_name(sheet)
            break
        except:
            print("This excel document or sheet is not available, please make sure that you \ntyped it correctly")
            return [None]
    result = []
    for pos in range(len(sheet.rows)):
        try:
            result.append(str(sheet.cell(row = pos + 1, column = 1).value))
            result.append(str(sheet.cell(row = pos + 1, column = 2).value))
        except:
            print("Something went wrong loading the excel file, see row {}".format(pos + 1))
            return [None]
    return result

def text_file(file_name):
    '''Reads from the given text file'''
    results = []
    try:
        file = open(file_name).read().splitlines()
    except:
        return [None]
    for line in file:
        line = line.strip().lower()
    if len(file[0]) == 1:
        results = file
        return results
    for line in file:
        for word in line.split():
            results.append(word)
    return results

def word_source()->list:
    '''Asks for a file name to read from and prints the results'''
    print("""Please enter the fileid of the file exactly as it is (including
the extension), please ensure that the file is present in the current
folder and that all words are listed in the first or second columns (excel)
or as a list seperated by spaces (text).\n""")
    results = [None]
    file = open('common_words.txt').read().splitlines()
    while (True):
        try:
            file = input("File name: ").strip()
            #print(file)
            #file = load_workbook(file_name)
            if file[-1] == 'x':
                results = excel_file(file)
            elif file[-1] == 't':
                results = text_file(file)
            if results != [None]:
                break
        except:
            print("That file name is not valid")
    print("\nFor future reference use this listing to help determine which definitions to use")
    print("Requested words as pairs: (ensure that this is correct before proceeding)")
    print("Word 1            Word 2")
    #print(len(results)/2)
    for i in range(int(len(results)/2)):
        print("{:18}{}".format(results[2*i], results[2*i + 1]))
    return results

def ask_for_word_defs(words: list)->list:
    '''Asks for the defintions to comapre each word to the other'''
    results = []
    synsets = []
    print("\nPlease wait, compiling word definitions...")
    for word in words:
        synsets.append(wordnet.synsets(word))
    for i in range(len(synsets)):
        if len(synsets[i]) == 0:
            print("\n{}\n".format('*'*80))
            print('No definitions for the word "{}" found'.format(words[i]))
            results.append(None)
            continue
        elif len(synsets[i]) == 1:
            print("\n{}\n".format('*'*80))
            print('Exactly 1 definition for the word "{}" found'.format(words[i]))
            results.append(synsets[i][0])
            print(results[-1].definition())
            continue
        print("\n{}\n".format('*'*80))
        print('Multiple definitions for the word "{}" detected. Please pick one by index #:'.format(words[i]))
        print('Defintions and parts of speech for "{}":'.format(words[i]))
        for def_index in range(len(synsets[i])):
          if synsets[i][def_index].pos() == 'n':
              print("{}: noun: {}".format(def_index,synsets[i][def_index].definition()))
          elif synsets[i][def_index].pos() == 'a':
              print("{}: adjective: {}".format(def_index,synsets[i][def_index].definition()))
          elif synsets[i][def_index].pos() == 's':
              print("{}: satellite adjective: {}".format(def_index,synsets[i][def_index].definition()))
          elif synsets[i][def_index].pos() == 'r':
              print("{}: adverb: {}".format(def_index,synsets[i][def_index].definition()))
          elif synsets[i][def_index].pos() == 'v':
              print("{}: verb: {}".format(def_index,synsets[i][def_index].definition()))
        while(True):
            try:
                index = int(input("Please enter the index of the desired definition for comparison: "))
                if index <= len(synsets[i]):
                    break
            except:
                print("That index is not valid")
            print("That index is out of range")
        results.append(synsets[i][index])
        print(results[-1].definition())
    return results

def similarity(words: list)->list:
    '''Calculates similarity based on the given synsets'''
    results = []
    synsets = ask_for_word_defs(words)
    print("\n{}\n".format('*'*80))
    for i in range(int(len(synsets)/2)):
        print("{:30}{}".format(synsets[2*i], synsets[2*i + 1]))
    print("\n{}\n".format('*'*80))
    print("Running comparisons...")
    for i in range(int(len(synsets)/2)):
        try:
            if (synsets[2*i] == None or synsets[2*i + 1] == None):
                results.append(["Undefined","Undefined", -1, -1, -1, "None", "None"])
                continue
        except:
            pass
        result = [words[2*i], words[2*i + 1], 0, 0, 0, synsets[2*i].definition(), synsets[2*i + 1].definition()]
        result[2] = wordnet.lch_similarity(synsets[2*i],synsets[2*i + 1])
        result[3] = wordnet.wup_similarity(synsets[2*i],synsets[2*i + 1])
        result[4] = wordnet.path_similarity(synsets[2*i],synsets[2*i + 1])
        results.append(result)
    print("\n{}\n".format('*'*80))
    return results

def handle_results(results)->None:
    '''Outputs the data as a file of specified name and type'''
    print("""Warning: for the moment this program cannot append to files,
only create and overwrite them, please be careful about using existing files
You can specify a different sheet name to use if you wish to add to a file

Also, please choose a file type of .xlsx (.xls is not currently supported)
and be sure that the file is not open in another window.
          """)
    while True:
        file_name = input("Please enter the name of the file you would to create, including the extension: ")
        print("\n{}\n".format(file_name))
        sure = input("Are you sure this is the file name you wish to use (y/n)").strip().lower()
        if sure not in ['y', 'yes']:
            continue
        if file_name.split('.')[1] != 'xlsx':
            print("This file type is not currently supported")
            continue
        sheet_name = input("Please enter a sheet name: ").strip()
        break
    excel.file_setup(results, file_name, sheet_name)
    print("\n{}\n".format('*'*80))
    print("Results have been saved!")
    return

if __name__ == '__main__':
    print('''Instructions:
Enter the file you would like to analyze, accepatble file types are '.txt' and
'.xlsx'. The program will assume white-space based formatting for text files
and 2 words, separated by column, per row for excel files. Once the words have
been read into the program it will list them as pairs. If the pairs are
incorrect you can reformat the input document and start over. If they are
correct you can wait until the word definitions are compiled into a list and
then for every word you entered the program will automattically display the
known definitions for each word. If the word has no definitions it will be
replaced by "None" and that pair will be skipped. If the word has exactly one
definition it will be entered automatically. If there are multiple definitions
you will be asked to pick one. After a defintion has been chosen it will be
printed back to you and the next word will be displayed. When all word
definitons are chosen the porgram will display the final synsets and ask you to
provide a name for a file to output the results to. The results are formatted
such that each word pair is listed followed by the three similarity comparisons
followed by the definitions used for the given words.\n''')
    words = word_source()
    results = similarity(words)
    handle_results(results)
