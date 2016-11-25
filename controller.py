# Controller
# Author :      Nathan Krueger
# Created       5:00 PM 7/16/15
# Last Updated  9:45 AM 11/24/16
# Version       2.7

import UI
import nltk
import xlrd
from openpyxl import *
from nltk.corpus import wordnet

#Debugging:
load_excel_files = False
load_NLTK_corpora = False
if not load_excel_files:
    print("""WARNING: You are in non-excel debugging mode,
this will result in data gathering failures and may result in a crash!""")

if load_NLTK_corpora:
    from nltk.book import *
else:
    print("""WARNING: You are in non-NLTK-corpus debugging mode,
this may cause program crashes when using mwac/swac!""")

#global variable declarations:
word = ''
corpus = 0
aoa = 'AoA.xlsx'
awl = 'AWL.xlsx'
subtlex = 'SUBTLEX.xlsx'
tasa = 'tasa.xlsx'
zeno = 'Zeno.xlsx'

#Note: excel coords act as follows:
#cell(1,1) -> A-1   and cell(2,11) -> B-11

def run()->None:
    """Initializes program"""
    #global corpus
    #may want to load excel files on a case by case basis later on
    print("""
Please cite as:
Lawrence, J., & Krueger, N. (2016). Bifrost: Bridging linguistic, cognitive and computer science resources (Version Master). Retrieved from https://github.com/Malthanatos/Bifrost

This program was developed with support from the University of California Academic Senate Council on Research, Computing, and Libraries (CORCL).

Please contract Joshua Lawrence at jflawren@uci.edu for any of the following papers that have used from data derived from the Bifrost program:

Lawrence, J. F., Hwang, J. K., Hagen, A., & Lin, G. (n.d.). What makes an academic word difficult to know?: Exploring lexical dimensions across novel measures of word knowledge. 

Lawrence, J.F., Hagen, A., Hwang, J. K., Lin, G., & Arne, L. (n.d.). Academic vocabulary and reading comprehension: Exploring the relationships across measures of vocabulary knowledge.

Lawrence, J. F., Lin, G., Jaeggi, S., Krueger, N., Hwang, J. K., & Hagen, A. (n.d.). Polysemy and semantic precision: Standardized semantic measures extracted from Wordnet for 100,000 words in English.

Lawrence, J. F., (n.d.) Semantic precision and polysemy: Key indices of word difficulty and utility for reading.
""")
    if load_excel_files:
        print("\nLoading excel files...")
        excel_setup()
        print("Done")
    UI.setup()
    print('''
Notes: some 2 part words can be analyzed, however, the results
       - of the analysis of such words may be inconsistant depending on
       - whether the input uses a space or an underscore to seperate them
       
       currently, when data from a corpus is loaded using mwac the related
       - words and dispersion plots will be displayed before any other
       - data such as definitions and statistics''')
    while True:
        interface_data = UI.interface()
        if interface_data[0] == 'quit':
            break
        elif interface_data == (None,None):
            continue
        UI.output_data(collect_data(interface_data))
    return

def collect_data(in_data)->list:
    """collects the data requested"""
    function = in_data[0]
    other = in_data[1]
    if function == 'newc':
        data = corupus_setup(other[0], other[1])
    if function == 'swa':
        data = analyze([other], 0)
    if function == 'swac':
        data = analyze([other[1]], 1, other[0])
    if function == 'mwaw':
        data = analyze(other, 2)
    if function == 'mwax':
        data = analyze(other, 0)
    if function == 'mwac':
        data = analyze(other[1], 1, other[0])
    if function == 'dwsa':
        data = dual_word(other)
    if function == 'polys':
        data = polysemy(other)
    if function == 'mindep':
        data = mindepth(other)
    if function == 'pol_min':
        data = polys_mindep(other)
    if function == 'dtree':
        data = depth_tree(other)
    if function == 'xhyper':
        data = xhyper(other)
    return (data, function)

def excel_setup()->None:
    """opens the necessary files/worksheets from excel documents"""
    global awl, aoa, tasa, subtlex, zeno
    #I tried to make a for loop for this but it never worked...
    try:
        awl = xlrd.open_workbook('AWL.xlsx')
        awl = awl.sheet_by_index(0)
        print("1/5")
    except:
        print("Failed to load file: AWL.xlsx")
    try:
        aoa = xlrd.open_workbook('AoA.xlsx')
        aoa = aoa.sheet_by_index(0)
        print("2/5")
    except:
        print("Failed to load file: AoA.xlsx")
    try:
        tasa = xlrd.open_workbook('tasa.xlsx')
        tasa = tasa.sheet_by_index(0)
        print("3/5")
    except:
        print("Failed to load file: tasa.xlsx")
    try:
        subtlex = xlrd.open_workbook('SUBTLEX.xlsx')
        subtlex = subtlex.sheet_by_index(0)
        print("4/5")
    except:
        print("Failed to load file: SUBTLEX.xlsx")

    try:
        zeno = xlrd.open_workbook('Zeno.xlsx')
        zeno = zeno.sheet_by_index(0)
        print("5/5")
    except:
        print("Failed to load file: Zeno.xlsx")
    return

def corupus_setup(file, name: str)->bool:
    '''sets up a new corpus to be selected from'''
    try:
        listing = open('corpora.txt', 'a')
        corpora = open('corpora.txt', 'r')
        index = int(corpora.read().splitlines()[-1].split()[0]) + 1
        listing.write('\n{}\t{}\t{}'.format(index,name,file))
        listing.close()
        corpora.close()
        UI.setup()
    except:
        print("Could not install new corpus...")
        return False
    return True

def pos_redef(pos: str)->str:
    '''converts the 1 letter synset POS into a real word'''
    if pos == 'n':
        return "noun"
    if pos == 'a':
        return "adjective"
    if pos == 's':
        return "satellite adjective"
    if pos == 'r':
        return "adverb"
    if pos == 'v':
        return "verb"

def analyze(words: [str], nltk_wn: int, corpus_id = 0)->[str]:
    """analyze a given word and report all available data"""
    global corpus
    print("\nGathering data...")
    if corpus_id != 0:
        if corpus_id < 10:
            corpus = eval('text' + str(corpus_id))
        else:
            corpus = UI.corpora[corpus_id-1].split('\t')[2]
            corpus = open(corpus, 'r').read()
    result = []
    x = 0
    for word in words:
        result.append(list())
        result[x].append(word)
        #nltk analysis
        if nltk_wn == 1:
            result[x].append(nltk_data(word))
        elif nltk_wn == 2:
            result[x] += word_def_data(word) #only works because python likes lists
            x += 1
            continue
        else:
            result[x].append(None)
        #wordnet analysis
        result[x].append(wordnet_data(word))
        #excel analysis
        result[x].append(excel_data(word))
        x += 1
    return result

def nltk_data(word: str)->[str]:
    """returns a list of semi-formatted nltk word data"""
    #similar words, occurences, frequency plot
    result = [None,None,None,None,None]
    #similar seems to auto-print and return None
    #technically part of UI:
    print('\nWords related to "{}" within selected corpus: '.format(word))
    try:
        corpus.similar(word)
    except:
        print("Can't print related words from non-NLTK corpora")
    result[0] = len(corpus)
    result[1] = len(set(corpus))
    result[2] = result[1] / result[0]
    result[3] = corpus.count(word)
    result[4] = 100 * result[3] / result[0]
    #see UI for more specifics here:
    try:
        print("\n{}'s distribution within the corpus: (see second window)".format(word))
        print("\nNote: in current development stage the program cannot continue until the \nfrequency distribution window is closed.")
        corpus.dispersion_plot([word])
    except (ValueError, ImportError):
        print("\nMatplot and/or numpy libraries are not installed, frequecny distribution cannot be displayed")
    except (AttributeError):
        print("Cannot display distribution plot for non-NLTK corpus")
    except(...):
        print("An unknown error occured while attempting to display the word's frequency distribution")
    return result

def wordnet_data(word: str)->[str]:
    """returns a list of semi-formatted wordnet word data"""
    #definitions, parts of speech, synonyms, antonyms, related words
    result = [[],[],[],set(),[],[]]
    word_info = wordnet.synsets(word)
    if (len(word_info) > 0):
        word_info = word_info[0]
    else:
        return result
    lemmas = wordnet.lemmas(word)
    for synset in wordnet.synsets(word):
        #if synset.name().split('.')[0] != word:
            #continue
        result[0].append(synset.definition())
        result[1].append(synset.pos())
    result[2] = word_info.lemma_names()
    for lemma in lemmas:
        for word in lemma.antonyms():
            result[3].add(word.name())
    result[4] = (word.name().split('.')[0] for word in word_info.hyponyms())
    result[5] = (word.name().split('.')[0] for word in word_info.hypernyms())
    return result

def word_def_data(word: str)->list:
    '''returns a list of word definitions, POS, synsets, etc'''
    #word, def #, synset, mindepth, POS, POS #, definition
    results = []
    word_info = wordnet.synsets(word)
    num = 0
    pos_num = 0
    pos = '?'
    for synset in word_info:
        new_pos = pos_redef(synset.pos())
        if pos != new_pos:
            pos = new_pos
            pos_num = 0
        else:
            pos_num += 1
        result = [word, num, synset, synset.min_depth(), pos, pos_num, synset.definition()]
        results.append(result)
        num += 1
    return results

def dual_word(words: [str])->list:
    '''returns a list of word similarities'''
    print("\nGathering data...")
    results = []
    similarities = []
    #can't say if word1 == None or ''
    on_even = False
    valid = True
    #functions take synsets
    #Need wordnet_ic.ic()
    #Currently, ic is WIP and requires corpus
    functions = [wordnet.lch_similarity, wordnet.wup_similarity, wordnet.path_similarity]
    #ic_functions = [wordnet.jcn_similarity, wordnet.lin_similarity, wordnet.res_similarity]
    for word in words:
        similarities.append(word)
        if not on_even:
            try:
                word1 = wordnet.synsets(word)[0]
                valid = True
            except:
                #Invalid word:
                valid = False
                continue
            on_even = True
        #word -> word2
        else:
            if not valid:
                similarities = [word,similarities[-1],0,0,0]
                results.append(similarities)
                similrities = []
                on_even = False
                valid = True
                continue
            try:
                word = wordnet.synsets(word)[0]
            except:
                #Invaid word, skip this
                on_even = False
                similarities = [similarities[0],similarities[1],0,0,0]
                results.append(similarities)
                similarities = []
                continue
            for f in functions:
                similarities.append(f(word1,word))
            #similarities.append(wordnet.lch_similarity(word1,word))
            on_even = False
            results.append(similarities)
            similarities = []
    return results

def polysemy(words: [str])->list:
    '''returns a list of polysemy data for a given word set'''
    '''if word_source == 'default':
        file = open('common words.txt')
        words = file.read().splitlines()
    elif word_source == 'manual':
        print("Please enter a string of words seperated only by spaces: ")
        words = input().strip().lower().split()
    #words = [w.lower() for w in word_list]'''
    print("\nGathering data...")
    result = []
    for word in words:
        word_data = [word,0,0,0,0,0,'N/A','N/A','N/A','N/A','N/A']
        word_info = wordnet.synsets(word)
        for synset in word_info:
            #if synset.name().split('.')[0] != word:
                #continue
            if synset.pos() == 'n':
                word_data[1] += 1
            if synset.pos() == 'a':
                word_data[2] += 1
            if synset.pos() == 's':
                word_data[3] += 1
            if synset.pos() == 'r':
                word_data[4] += 1
            if synset.pos() == 'v':
                word_data[5] += 1
        result.append(word_data)
    return result

def mindepth(words: [str])->list:
    '''returns a list of tuples of a word and its min depth'''
    print("\nGathering data...")
    result = []
    for word in words:
        word_data = [word,'N/A','N/A','N/A','N/A','N/A',-1,-1,-1,-1,-1]
        word_info = wordnet.synsets(word)
        for index in range(len(word_info)):
            #if word_info[index].name().split('.')[0] != word:
                #continue
            if word_info[index].pos() == 'n' and word_data[6] == -1:
                word_data[6] = word_info[index].min_depth()
            if word_info[index].pos() == 'a' and word_data[7] == -1:
                word_data[7] = word_info[index].min_depth()
            if word_info[index].pos() == 's' and word_data[8] == -1:
                word_data[8] = word_info[index].min_depth()
            if word_info[index].pos() == 'r' and word_data[9] == -1:
                word_data[9] = word_info[index].min_depth()
            if word_info[index].pos() == 'v' and word_data[10] == -1:
                word_data[10] = word_info[index].min_depth()
        result.append(word_data)
    return result

def polys_mindep(words: [str])->list:
    '''returns a list of lists of words and their depth and polys'''
    #I could shorten this by calling both and merging them, but calling synsets is expensive
    #  and I don't want to do it twice if I can help it
    print("\nGathering data...")
    result = []
    for word in words:
        word_data = [word,0,0,0,0,0,-1,-1,-1,-1,-1]
        word_info = wordnet.synsets(word)
        for index in range(len(word_info)):
            #if word_info[index].name().split('.')[0] != word:
                #continue
            if word_info[index].pos() == 'n':
                if word_data[6] == -1:
                    word_data[6] = word_info[index].min_depth()
                word_data[1] += 1
            if word_info[index].pos() == 'a':
                if word_data[7] == -1:
                    word_data[7] = word_info[index].min_depth()
                word_data[2] += 1
            if word_info[index].pos() == 's':
                if word_data[8] == -1:
                    word_data[8] = word_info[index].min_depth()
                word_data[3] += 1
            if word_info[index].pos() == 'r':
                if word_data[9] == -1:
                    word_data[9] = word_info[index].min_depth()
                word_data[4] += 1
            if word_info[index].pos() == 'v':
                if word_data[10] == -1:
                    word_data[10] = word_info[index].min_depth()
                word_data[5] += 1
        result.append(word_data)
    return result

def depth_tree(word)->str:
    '''returns the word's depth tree'''
    print("Note: only nouns have dtrees, so only noun defintions are displayed")
    print("\nGathering data...")
    #word, defintions, pos, dtrees
    result = ['',[],[],[]]
    result[0] = word
    word_info = wordnet.synsets(word)
    hyp = lambda w: w.hypernyms()
    if (len(word_info) > 0):
        word_info = word_info[0]
    else:
        return result
    #iterating over word_info does not work:
    for synset in wordnet.synsets(word):
        result[1].append(synset.definition())
        result[2].append(synset.pos())
        result[3].append(synset.tree(hyp))
    return result
    #word = wordnet.synsets(word)[0]
    #hyp = lambda w:w.hypernyms()
    #return word.tree(hyp)

def valueAt(pos, L)->str:
    try:
        if pos == 0:
            return L[0]
        else:
            return valueAt(pos - 1, L[1])
    except:
        return None

def xhyper(words)->[str]:
    '''returns the highest order x hypernyms'''
    x = UI.request_x()
    print("\nNote: this program will use of the first word definition it finds and the\nfirst parallel synset if there are any")
    print("\nGathering data...")
    result = [x]
    hyp = lambda w: w.hypernyms()
    depth = lambda L: isinstance(L, list) and max(map(depth, L))+1
    for i in range(len(words)):
        synsets = wordnet.synsets(words[i])
        if len(synsets) > 0:
            hyper = wordnet.synsets(words[i])[0].tree(hyp)
            d = depth(hyper) - 1
            xhyper = []
            for j in range(x):
                xhyper.append(valueAt(d - j, hyper))
            result.append([words[i], hyper[0], xhyper])
        else:
            result.append([words[i], [None]])
    return result

#Note: excel data is not optimally formatted, so linear search is used
#      this may be changed to binary search later on when the files are in alphabetical order

def excel_data(word: str)->[str]:
    """returns a list of all excel data"""
    result = [None,None,None,None,None]
    result[0] = 0
    result[1] = [None,None,None,None,None,None]
    result[2] = 0
    result[3] = [None,None,None,None,None,None,None,None]
    result[4] = [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None]
    try:
        result[0] = awl_data(word)
    except:
        print('\nLoading AWL data for word: "{}" failed\n'.format(word))
    try:
        result[1] = aoa_data(word)
    except:
        print('\nLoading AoA data for word: "{}" failed\n'.format(word))
    try:
        result[2] = tasa_data(word)
    except:
        print('\nLoading TASA data for word: "{}" failed\n'.format(word))
    try:
        result[3] = subtlex_data(word)
    except:
        print('\nLoading SUBLTEX data for word: "{}" failed\n'.format(word))
    try:
        result[4] = zeno_data(word)
    except:
        print('\nLoading Zeno data for word: "{}" failed\n'.format(word))
    return result    

def awl_data(word: str)->int:
    """returns all available awl data"""
    result = 0
    for pos in range(awl.nrows):
        if (word == awl.cell(pos,0).value):
            result = awl.cell(pos,1).value   #AWL rating
            break
        pos += 1
    return result

def aoa_data(word: str)->[str]:
    """returns all available aoa data"""
    result = [None,None,None,None,None,None]
    for pos in range(aoa.nrows):
        if (word == aoa.cell(pos,0).value):
            result[0] = aoa.cell(pos,1).value   #OccurTotal
            result[1] = aoa.cell(pos,2).value   #OccuurNum
            result[2] = aoa.cell(pos,3).value   #Freq_pm
            result[3] = aoa.cell(pos,4).value   #Rating.Mean
            result[4] = aoa.cell(pos,5).value   #Rating.SD
            result[5] = aoa.cell(pos,6).value   #Dunno
            break
        pos += 1
    return result

def tasa_data(word: str)->int:
    """locates and returns a word's tasa data"""
    result = 0
    for pos in range(tasa.nrows):
        if (word == tasa.cell(pos,0).value):
            result = tasa.cell(pos,1).value
            break
        pos += 1
    return result

def subtlex_data(word: str)->[str]:
    """returns all available subtlex data"""
    result = [None,None,None,None,None,None,None,None]
    for pos in range(subtlex.nrows):
        if (word == subtlex.cell(pos,0).value):
            result[0] = subtlex.cell(pos,1).value   #FREQcount
            result[1] = subtlex.cell(pos,2).value   #CScount
            result[2] = subtlex.cell(pos,3).value   #FREQlow
            result[3] = subtlex.cell(pos,4).value   #Cdlow
            result[4] = subtlex.cell(pos,5).value   #SUBTL_WF
            result[5] = subtlex.cell(pos,6).value   #Log_10(WF)
            result[6] = subtlex.cell(pos,7).value   #SUBTL_CD
            result[7] = subtlex.cell(pos,8).value   #Log_10(CD)
            break
        pos += 1
    return result

def zeno_data(word: str)->[str]:
    """returns all available zeno data"""
    result = [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None]
    for pos in range(zeno.nrows):
        if (word == zeno.cell(pos,0).value):
            result[0] = zeno.cell(pos,1).value   #sfi
            result[1] = zeno.cell(pos,2).value   #d
            result[2] = zeno.cell(pos,3).value   #u
            result[3] = zeno.cell(pos,4).value   #f
            result[4] = zeno.cell(pos,5).value   #gr1
            result[5] = zeno.cell(pos,6).value   #gr2
            result[6] = zeno.cell(pos,7).value   #gr3
            result[7] = zeno.cell(pos,8).value   #gr4
            result[8] = zeno.cell(pos,9).value   #gr5
            result[9] = zeno.cell(pos,10).value   #gr6
            result[10] = zeno.cell(pos,11).value  #gr7
            result[11] = zeno.cell(pos,12).value  #gr8
            result[12] = zeno.cell(pos,13).value  #gr9
            result[13] = zeno.cell(pos,14).value  #gr10
            result[14] = zeno.cell(pos,15).value  #gr11
            result[15] = zeno.cell(pos,16).value  #gr12
            result[16] = zeno.cell(pos,17).value  #gr13
            break
        pos += 1
    return result

if __name__ == '__main__':
    run()
