from openpyxl import *

wb = load_workbook('simple.xlsx')
s = wb.get_sheet_by_name('Sheet1')

s.cell(row = 1,column = 1).value = 'light'
wb.save('simple.xlsx')


'''print('Sheet:',s.title)
#print(s.rows)
#print(s.columns)
values = []
index = 0
for row in s.rows:
    #print(row)
    values.append(list())
    for cell in row:
        if cell.value != None:
            values[index].append(cell.value)
    index += 1
        #print(cell.value)
print(values)'''
'''
for row in s.rows:
    values = []
    #for cell in row:
    values.append(row.value)
    print(','.join(str(values)))
print()'''

#Functions:
'''
wb = load_workbook('name', read_only = True/False)
wb.active() returns the first sheet
wb.craete_sheet() or (#) create sheet at index
wb.get_sheet_names()
ws.get_sheet_by_name('name')
ws.title('title')
wb.get_sheet_by_name('name')
wb.sheet_names()
ws['A4'] returns cell A4
ws.cell(4,2) returns cell at 4,2
c.value = 'hello'
wb.save('name')
'''
