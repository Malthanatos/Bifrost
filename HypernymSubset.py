# HypernymSubset
# Author :      Nathan Krueger
# Created       10:30 AM 1/5/17
# Last Updated  1:35 PM 5/6/17
# Version       0.76

from openpyxl import *
import excel

#file format:
#Word Synset used POS Hypernym 1, ..., Hypernym 20
#1    2           3   4                23

#Note: Hypernym 1 us most general, if x is the depth then Hypernym x is the
#      synset that is being used

#xhper sheet allows words that return as None, as in, not found
#Both sheets always end with a value of None for each row

MAX_COL = 23
#1 -> allow not found words, 0 -> all words that had valid synsets, -1 -> test
ALLOW_NOT_FOUND = 0

file_name = "100k_xhyper.xlsx"
sheets = ["xhyper2", "xhyper", "xhyper3"]
counts = [146494, 198120, 47]

sheet_name = sheets[ALLOW_NOT_FOUND]
count = counts[ALLOW_NOT_FOUND]

#test: Hypernym 3 == Synset('object.n.01')
#returns: Synset('light.n.02'), Synset('light.n.05'), Synset('lighter.n.02')
#Result format: light | Synset('light.n.02') | Synset('object.n.01') | x
#The original word, the synset used, the synset searched for, and the depth at which it was found/searched for

def UI()->(str, int):
    '''Asks user for a specific hypernym and depth'''
    print("Note: Synsets are composed of 3 elements, a base word, an abreviated part of\nspeech, and a 2 digit ID number to differentaite between identical base words\nwith different meanings\n")
    print("Note: Parts of speech are abreviated as follows: noun -> n, verb -> v,\nadjective -> a, adverb -> r, satellite adjective -> s\n")
    print("Note: synsets typically do not have ID numbers greater than 10, generally more\nspecific synsets will have fewer IDs, the word 'light', as a counter-example, is\nvery general (it has 47 definitions) and as such one of its synsets' base words ('light') has a maximum synset ID of 24\n")
    hypernym = ""
    split_hypernym = ["", "", ""]
    depth = -1
    while (True):
        if (hypernym == ""):
            hypernym = input("Please enter a hypernym to find including the word, part of speech, and synset\nID number (ex. object.n.01): ").strip().lower()
            split_hypernym = hypernym.split('.')
            if (len(split_hypernym) != 3):
                print("Error: {} is not a valid synset format".format(hypernym))
                hypernym = ""
            elif (split_hypernym[1] not in ['n', 'a', 's', 'r', 'v']):
                print("Error: the part of speech you entered, {}, was not valid".format(hypernym.split('.')[1]))
                hypernym = ""
            elif (len(split_hypernym[2]) != 2):
                print("Error: the synset ID must be exactly 2 digits, even if the first digit is 0")
                hypernym = ""
            try:
                int(split_hypernym[2])
            except:
                hypernym = ""
                print("Error: synset ID must be an integer")
            continue
        if (depth == -1):
            depth = input("Please enter a depth to search (0 for all depths, 1 is most general, high\ndepths may not return useful results, the maximum depth for all words is 20): ").strip()
            try:
                depth = int(depth)
                if (depth < 0 or depth > 20):
                    print("Error: depth must be in the range 0 - 20, inclusive")
                    depth = -1
                    continue
                break
            except:
                print("Error: {} is not a valid integer".format(depth))
                depth = -1
                continue
    return (hypernym, depth)

def search(sheet, hypernym: str, at_depth: int)->[str]:
    '''Locates any matching Synsets in the file using depth_at to determine which column to search'''
    results = []
    for pos in range(count):
        try:
            current = str(sheet.cell(row = pos + 1, column = at_depth + 3).value)
            if (current == "None"):
                continue
            Synset = "Synset('{}')".format(hypernym)
            if (current == Synset):
                word = str(sheet.cell(row = pos + 1, column = 1).value)
                base_synset = str(sheet.cell(row = pos + 1, column = 2).value)
                results.append([word, base_synset, Synset, at_depth])
        except:
            continue
    return results

def fetch(to_search: (str, int))->[str]:
    '''Fetches all words that have the hypernym that UI got at the given depth, or all depths if given depth is 0'''
    hypernym, depth = to_search
    results = [["Fetching:", hypernym, "at depth:", depth]]
    print("Searching for: {}, at {}".format(hypernym, "depth: " + str(depth) if depth != 0 else "any depth"))
    try:
        file = load_workbook(file_name)
        sheet = file.get_sheet_by_name(sheet_name)
    except:
        print("Error: excel file, '{}', or sheet '{}', not found, be sure they are not open and are in the same folder as this Python file".format(file_name, sheet_name))
    if (depth != 0):
        return results + search(sheet, hypernym, depth)
    #Search all depths
    for x in range(20):
        results += search(sheet, hypernym, x + 1)
    return results
    
def output(data: [str]):
    '''sends all data from run to an excel file named after the defined hypernym'''
    out_file_name = "HypernymSubset using {} at {}.xlsx".format(data[0][1], ("depth " + str(data[0][3])) if data[0][3] != 0 else "any depth")
    print("Done.\nSaving file as: {}".format(out_file_name))
    excel.file_setup([data, 'hypersub'], out_file_name, "HypernymSubset")
    return

if(__name__ =='__main__'):
    output(fetch(UI()))
