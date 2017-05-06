# SynsetFinder
# Author :      Nathan Krueger
# Created       12:50 AM 1/7/17
# Last Updated  1:00 AM 1/7/17
# Version       0.25 ("functional")

from openpyxl import *

file_name = "nouns.xlsx"
sheet_name = "nouns"

def UI()->str:
    '''Asks user for a file containing the search synsets'''
    while (True):
        in_file = input("Please enter a input file with synsets to search (.xlsx only): ").strip().lower()
        if (in_file.split(".")[-1] != "xlsx"):
            print("Error: file type not supported")
            continue
        else:
            break
    return in_file

def search(sheet, synset: str)->[str]:
    '''Locates any matching Synsets in the file using depth_at to determine which column to search'''
    '''results = []
    for pos in range(count):
        try:
            current = str(sheet.cell(row = pos + 1, column = 12).value)
            if (current == synset):
                word = str(sheet.cell(row = pos + 1, column = 1).value)
                base_synset = str(sheet.cell(row = pos + 1, column = 2).value)
                results.append([word, base_synset, synset, at_depth])
        except:
            continue
    return results'''

"""def load(to_search: str):
    '''Loads the synsets to look for from the UI'''
    result = []
    try:
        file = load_workbook(file_name)
        sheet = file.get_sheet_by_name(sheet_name)
    except:
        print("Error: excel file, '{}', or sheet '{}', not found, be sure they are not open and are in the same folder as this Python file".format(file_name, sheet_name))
        return []
    for x in range(35586):
        try:
            current = str(sheet.cell(row = x + 1, column = 12).value)
            if (current != to_search):
        except:
            continue
        result.append([])
        for y in range(70):
            result[-1].append(str(sheet.cell(row = x + 1, column = y + 1).value))
    print(result)
    return result"""

def load(to_search: str):
    '''Loads the synsets to look for from the UI'''
    result = []
    try:
        file = load_workbook(to_search)
        sheet = file.worksheets[0]
    except:
        print("Error: excel file, '{}', not found, be sure they are not open and are in the same folder as this Python file".format(to_search))
        return []
    for x in range(len(tuple(sheet.rows))):
        result.append(str(sheet.cell(row = x + 1, column = 1).value))
    return result

def fetch(to_search: str)->[str]:
    '''Fetches all words that have the hypernym that UI got at the given depth, or all depths if given depth is 0'''
    results = []
    search = load(to_search)
    try:
        file = load_workbook(file_name)
        sheet = file.get_sheet_by_name(sheet_name)
    except:
        print("Error: excel file, '{}', or sheet '{}', not found, be sure they are not open and are in the same folder as this Python file".format(file_name, sheet_name))
    for x in range(35585):
        current = str(sheet.cell(row = x + 1, column = 12).value)
        for i in range(len(search)):
            if (current == search[i]):
                #print(current)
                del search[i]
                #print(len(search))
                results.append([])
                for y in range(70):
                    results[-1].append(str(sheet.cell(row = x + 1, column = y + 1).value))
                break
    #print(results)
    return (results, to_search)
    
def output(data: ([[str]], str)):
    '''sends all data from run to an excel file named after the defined hypernym'''
    data, to_search = data
    out_file_name = "merged {}".format(to_search)
    print("Done.\nSaving file as: {}".format(out_file_name))
    #print(data)
    try:
        file = Workbook()
        sheet = file.active
        sheet.title = sheet_name
    except:
        print("Err: make sure that the file is closed if it exists!")
        file_setup(data, file_name, sheet_name)
    for x in range(len(data)):
        for y in range(70):
            sheet.cell(row = x + 1, column = y + 1).value = data[x][y]
    file.save(out_file_name)
    return

if(__name__ =='__main__'):
    output(fetch(UI()))
