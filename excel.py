# Excel
# Author :      Nathan Krueger
# Created       11:45 AM 8/9/15
# Last Updated  4:00 AM 12/3/16
# Version       2.95

from openpyxl import *

file = None
sheet = None

def file_setup(value, file_name: str, sheet_name = '', sheet_index = 0)->None:
    '''sets up the files header row'''
    data, function = value
    if type(data) != list or (type(data[0]) != int) and (type(data[0]) != list or len(data[0]) == 0):
        print("\n\n\nSomething happened to the data... Either you screwed up or I did. Double check and if you think I did then this is a bug, write down what you did and send it to me.\n\n\n")
        return
    global file, sheet
    try: #file exists, add a new sheet
        file = load_workbook(file_name)
        sheet = file.create_sheet(title = sheet_name)
    except: #file does not exist, rename first sheet
        try:
            file = Workbook()
            sheet = file.active
            sheet.title = sheet_name
        except:
            print("Err: make sure that the file is closed if it exists!")
            file_setup(data, file_name, sheet_name)
    if function == 'pol_min':
        polys_mindep_setup(data)
    elif function == 'dwsa':
        similarity(data)
    elif function == 'bssa':
        similarity_BSSA(data)
    elif function == 'swa' or function == 'mwa':
        word_def_setup(data)
    elif function == 'xhyper':
        xhyper_setup(data)
    else:
        excel_based_setup(data)
    file.save(file_name)
    return

def word_def_setup(data)->None:
    '''sets up a word definition data spreadsheet'''
    sheet.cell(row = 1, column = 1).value = 'Word'
    sheet.cell(row = 1, column = 2).value = 'Synset #'
    sheet.cell(row = 1, column = 3).value = 'Synset'
    sheet.cell(row = 1, column = 4).value = 'Depth'
    sheet.cell(row = 1, column = 5).value = 'POS'
    sheet.cell(row = 1, column = 6).value = 'POS #'
    sheet.cell(row = 1, column = 7).value = 'Definition'
    write_word_def(data)
    return

def polys_mindep_setup(data)->None:
    '''sets up a polys data spreadsheet'''
    sheet.cell(row = 1, column = 1).value = 'Word'
    sheet.cell(row = 1, column = 2).value = 'polys.noun'
    sheet.cell(row = 1, column = 3).value = 'polys.adj'
    sheet.cell(row = 1, column = 4).value = 'polys.sat_adj'
    sheet.cell(row = 1, column = 5).value = 'polys.adv'
    sheet.cell(row = 1, column = 6).value = 'polys.verb'
    sheet.cell(row = 1, column = 7).value = 'mindep.noun'
    sheet.cell(row = 1, column = 8).value = 'mindep.adj'
    sheet.cell(row = 1, column = 9).value = 'mindep.sat_adj'
    sheet.cell(row = 1, column = 10).value = 'mindep.adv'
    sheet.cell(row = 1, column = 11).value = 'mindep.verb'
    write_polys_mindep(data)
    return

def excel_based_setup(data)->None:
    '''sets up a spreadsheet based on excel data'''
    sheet.cell(row = 1, column = 1).value = 'Word'
    sheet.cell(row = 1, column = 2).value = 'Rating.Mean'
    sheet.cell(row = 1, column = 3).value = 'Rating.SD'
    sheet.cell(row = 1, column = 4).value = 'SUBTL_WF'
    sheet.cell(row = 1, column = 5).value = 'Log_10(WF)'
    sheet.cell(row = 1, column = 6).value = 'SUBTL_CD'
    sheet.cell(row = 1, column = 7).value = 'Log_10(CD)'
    sheet.cell(row = 1, column = 8).value = 'Zeno.sfi'
    sheet.cell(row = 1, column = 9).value = 'Zeno.d'
    write_excel(data)
    return

def similarity(data)->None:
    ''''sets up existing spreadsheet of word pairs to '''
    sheet.cell(row = 1, column = 1).value = 'Word 1'
    sheet.cell(row = 1, column = 2).value = 'Word 2'
    sheet.cell(row = 1, column = 3).value = 'LCH'
    sheet.cell(row = 1, column = 4).value = 'WUP'
    sheet.cell(row = 1, column = 5).value = 'Path'
    #sheet.cell(row = 1, column = 6).value = 'JCN'
    #sheet.cell(row = 1, column = 7).value = 'LIN'
    #sheet.cell(row = 1, column = 8).value = 'Res'
    write_sim(data)
    return

def similarity_BSSA(data)->None:
    ''''sets up existing spreadsheet of word pairs to '''
    sheet.cell(row = 1, column = 1).value = 'Word 1'
    sheet.cell(row = 1, column = 2).value = 'Word 2'
    sheet.cell(row = 1, column = 3).value = 'LCH'
    sheet.cell(row = 1, column = 4).value = 'WUP'
    sheet.cell(row = 1, column = 5).value = 'Path'
    sheet.cell(row = 1, column = 6).value = 'Word 1 Definition'
    sheet.cell(row = 1, column = 7).value = 'Word 2 Definition'
    write_sim_BSSA(data)
    return

def xhyper_setup(data)->None:
    ''''sets up existing spreadsheet of word pairs to '''
    sheet.cell(row = 1, column = 1).value = 'Word'
    sheet.cell(row = 1, column = 2).value = 'Synset used'
    sheet.cell(row = 1, column = 3).value = 'POS'
    for x in range(data[0]):
        sheet.cell(row = 1, column = x + 4).value = 'Hypernym {}'.format(x + 1)
    write_xhyper(data)
    return

def write_excel(data)->None:
    '''writes excel data to an excel file'''
    index = 2
    for word in data:
        sheet.cell(row = index, column = 1).value = word[0]
        sheet.cell(row = index, column = 2).value = word[3][1][3]
        sheet.cell(row = index, column = 3).value = word[3][1][4]
        sheet.cell(row = index, column = 4).value = word[3][3][4]
        sheet.cell(row = index, column = 5).value = word[3][3][5]
        sheet.cell(row = index, column = 6).value = word[3][3][6]
        sheet.cell(row = index, column = 7).value = word[3][3][7]
        sheet.cell(row = index, column = 8).value = word[3][4][0]
        sheet.cell(row = index, column = 9).value = word[3][4][1]
        index += 1
    return

def write_word_def(data)->None:
    '''writes excel data to an excel file'''
    index = 2
    for word in data:
        for synset in word[1:]:
            sheet.cell(row = index, column = 1).value = synset[0]
            sheet.cell(row = index, column = 2).value = synset[1]
            sheet.cell(row = index, column = 3).value = str(synset[2])
            sheet.cell(row = index, column = 4).value = synset[3]
            sheet.cell(row = index, column = 5).value = synset[4]
            sheet.cell(row = index, column = 6).value = synset[5]
            sheet.cell(row = index, column = 7).value = synset[6]
            index += 1
    return

def write_polys_mindep(data)->None:
    '''writes polysemy data to the file'''
    index = 2
    for word in data:
        sheet.cell(row = index, column = 1).value = word[0]
        sheet.cell(row = index, column = 2).value = word[1]
        sheet.cell(row = index, column = 3).value = word[2]
        sheet.cell(row = index, column = 4).value = word[3]
        sheet.cell(row = index, column = 5).value = word[4]
        sheet.cell(row = index, column = 6).value = word[5]
        sheet.cell(row = index, column = 7).value = word[6]
        sheet.cell(row = index, column = 8).value = word[7]
        sheet.cell(row = index, column = 9).value = word[8]
        sheet.cell(row = index, column = 10).value = word[9]
        sheet.cell(row = index, column = 11).value = word[10]
        index += 1
    return

def write_sim(data)->None:
    '''writes similarity data to file'''
    index = 2
    for word in data:
        sheet.cell(row = index, column = 1).value = word[0]
        sheet.cell(row = index, column = 2).value = word[1]
        sheet.cell(row = index, column = 3).value = word[2]
        sheet.cell(row = index, column = 4).value = word[3]
        sheet.cell(row = index, column = 5).value = word[4]
        #sheet.cell(row = index, column = 6).value = word[5]
        #sheet.cell(row = index, column = 7).value = word[6]
        #sheet.cell(row = index, column = 8).value = word[7]
        index +=1
    return

def write_sim_BSSA(data)->None:
    '''writes BSSA similarity data to file'''
    index = 2
    for word in data:
        sheet.cell(row = index, column = 1).value = word[0]
        sheet.cell(row = index, column = 2).value = word[1]
        sheet.cell(row = index, column = 3).value = word[2]
        sheet.cell(row = index, column = 4).value = word[3]
        sheet.cell(row = index, column = 5).value = word[4]
        sheet.cell(row = index, column = 6).value = word[5]
        sheet.cell(row = index, column = 7).value = word[6]
        index +=1
    return

def write_xhyper(data)->None:
    '''writes BSSA similarity data to file'''
    index = 2
    for word in data[1:]:
        sheet.cell(row = index, column = 1).value = word[0]
        sheet.cell(row = index, column = 2).value = str(word[2])
        sheet.cell(row = index, column = 3).value = word[1]
        for x in range(data[0]):
            try:
                sheet.cell(row = index, column = x + 4).value = str(word[3][x])
            except:
                break
        index +=1
    return
